"""REST API routes."""

import re

from fastapi import APIRouter, Depends, HTTPException, Request, Response
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from .. import models
from ..auth import audit, check_rate_limit, clear_failures, get_current_user, record_failure, verify_password
from .. import allowlist
from ..bridge import manager
from ..config import get_settings
from ..db import get_session
from ..ha.state import cache
from ..schemas import (
    EntityOut,
    HealthOut,
    LayoutIn,
    LayoutOut,
    PlacementIn,
    PlacementOut,
    ServiceCall,
)

router = APIRouter(prefix="/api")
protected = APIRouter(prefix="/api", dependencies=[Depends(get_current_user)])



# -- health ------------------------------------------------------------------
@router.get("/health", response_model=HealthOut)
async def health() -> HealthOut:
    settings = get_settings()
    return HealthOut(
        status="ok",
        mode=manager.mode,
        ha_connected=cache.ha_connected,
        entity_count=len(cache),
    )


# -- entities ----------------------------------------------------------------
@protected.get("/entities", response_model=list[EntityOut])
async def list_entities() -> list[EntityOut]:
    return [EntityOut(**e.to_dict()) for e in cache.snapshot()]


@protected.get("/entities/{entity_id}", response_model=EntityOut)
async def get_entity(entity_id: str) -> EntityOut:
    ent = cache.get(entity_id)
    if not ent:
        raise HTTPException(404, f"unknown entity: {entity_id}")
    return EntityOut(**ent.to_dict())


# -- service calls -----------------------------------------------------------
@protected.post("/services/{domain}/{service}", status_code=202)
async def call_service(
    domain: str,
    service: str,
    body: ServiceCall,
    request: Request,
    user: models.User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> dict:
    if not allowlist.is_allowed(domain, service):
        raise HTTPException(403, f"service {domain}.{service} is not exposed by this app")
    if not cache.get(body.entity_id):
        raise HTTPException(404, f"unknown entity: {body.entity_id}")

    # The PIN is an app-level field: pop it unconditionally so it can never
    # be forwarded to HA in any service payload.
    pin = body.data.pop("pin", None)

    # Arm/disarm PIN gate. Enforced here (not in the UI) so the API itself
    # is protected; per-user, so the audit trail says WHO armed/disarmed.
    # Users without a PIN configured are not gated (set PINs for every
    # account that can reach a wall panel).
    if domain == "alarm_control_panel" and user.pin_hash is not None:
        ip = request.client.host if request.client else "?"
        check_rate_limit(f"pin:{user.username}", ip)
        if not pin:
            raise HTTPException(403, "pin_required")
        if not verify_password(str(pin), user.pin_hash):
            record_failure(f"pin:{user.username}", ip)
            await audit(session, user.username, "alarm_pin_fail", f"{service} -> {body.entity_id}")
            raise HTTPException(403, "pin_invalid")
        clear_failures(f"pin:{user.username}", ip)

    bridge = manager.bridge
    if bridge is None:
        raise HTTPException(503, "bridge not running")
    try:
        await bridge.call_service(domain, service, body.entity_id, body.data)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(502, f"service call failed: {exc}") from exc
    await audit(session, user.username, "service_call", f"{domain}.{service} -> {body.entity_id}")
    return {"ok": True}


# -- sensor export -----------------------------------------------------------
# Spreadsheet registry of every entity the bridge sees, joined with saved
# placements. Two sheets: "Sensors" (the security/environment devices, with
# room/floor/coords — the pairing tracker) and "All entities" (raw dump for
# YAML/automation work, exact entity_ids included). Built in-memory with
# openpyxl; nothing is written to disk.
SENSOR_DOMAINS = {"binary_sensor", "lock", "siren", "climate", "valve"}


@protected.get("/export/sensors.xlsx")
async def export_sensors(session: AsyncSession = Depends(get_session)) -> Response:
    from datetime import datetime as _dt
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    result = await session.execute(select(models.SensorPlacement))
    place = {p.entity_id: p for p in result.scalars()}
    ents = sorted(cache.snapshot(), key=lambda e: (e.domain, e.entity_id))
    floor_name = {0: "Ground", 1: "Upstairs"}

    wb = Workbook()
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="1F2937")

    def sheet(ws, headers, rows):
        ws.append(headers)
        for c in ws[1]:
            c.font = head_font
            c.fill = head_fill
        for r in rows:
            ws.append(r)
        for i, h in enumerate(headers, 1):
            width = max([len(str(h))] + [len(str(row[i - 1])) for row in rows] or [10])
            ws.column_dimensions[get_column_letter(i)].width = min(width + 2, 48)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    ws1 = wb.active
    ws1.title = "Sensors"
    sensors = [e for e in ents if e.domain in SENSOR_DOMAINS]
    sheet(
        ws1,
        ["Entity ID", "Friendly Name", "Type", "Room", "Floor", "Placed", "X", "Y", "State", "Battery %", "Last Changed"],
        [
            [
                e.entity_id,
                e.friendly_name,
                str(e.attributes.get("device_class") or e.domain),
                place[e.entity_id].room if e.entity_id in place else "",
                floor_name.get(place[e.entity_id].floor, place[e.entity_id].floor) if e.entity_id in place else "",
                "yes" if e.entity_id in place else "NOT PLACED",
                round(place[e.entity_id].x, 2) if e.entity_id in place else "",
                round(place[e.entity_id].y, 2) if e.entity_id in place else "",
                e.state,
                e.attributes.get("battery", ""),
                e.last_changed.isoformat(sep=" ", timespec="seconds"),
            ]
            for e in sensors
        ],
    )

    ws2 = wb.create_sheet("All entities")
    sheet(
        ws2,
        ["Entity ID", "Friendly Name", "Domain", "Device Class", "State", "Last Changed"],
        [
            [
                e.entity_id,
                e.friendly_name,
                e.domain,
                str(e.attributes.get("device_class") or ""),
                e.state,
                e.last_changed.isoformat(sep=" ", timespec="seconds"),
            ]
            for e in ents
        ],
    )

    buf = BytesIO()
    wb.save(buf)
    fname = f"homehub-sensors-{_dt.now().strftime('%Y%m%d')}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# -- weather radar (RainViewer proxy) ----------------------------------------
# RainViewer is an explicitly approved external dependency (see CLAUDE.md).
# Proxied through the backend so clients — wall panels on default-deny
# VLANs — only ever talk to this app. No API key involved. Meta is cached
# for 60s; tiles in a small in-memory LRU (they're immutable per-timestamp).
import time as _time
from collections import OrderedDict

_radar_meta: dict = {"at": 0.0, "data": None}
_radar_tiles: OrderedDict[str, bytes] = OrderedDict()
_RADAR_TILE_CACHE = 400


async def _radar_settings(session: AsyncSession) -> tuple[float | None, float | None]:
    lat = await session.get(models.AppSetting, "latitude")
    lon = await session.get(models.AppSetting, "longitude")
    try:
        return (float(lat.value) if lat and lat.value else None,
                float(lon.value) if lon and lon.value else None)
    except ValueError:
        return None, None


@protected.get("/radar/meta")
async def radar_meta(session: AsyncSession = Depends(get_session)) -> dict:
    import httpx

    lat, lon = await _radar_settings(session)
    now = _time.monotonic()
    if _radar_meta["data"] is None or now - _radar_meta["at"] > 60:
        try:
            async with httpx.AsyncClient(timeout=8) as client:
                r = await client.get("https://api.rainviewer.com/public/weather-maps.json")
                r.raise_for_status()
                _radar_meta.update(at=now, data=r.json())
        except Exception as exc:  # noqa: BLE001
            if _radar_meta["data"] is None:
                raise HTTPException(503, f"radar service unreachable: {exc}") from exc
    data = _radar_meta["data"]
    frames = [
        {"ts": f["time"], "path": f["path"], "nowcast": False}
        for f in data.get("radar", {}).get("past", [])
    ] + [
        {"ts": f["time"], "path": f["path"], "nowcast": True}
        for f in data.get("radar", {}).get("nowcast", [])
    ]
    return {"frames": frames, "lat": lat, "lon": lon}


_RADAR_PATH_RE = re.compile(r"^/v2/radar/[A-Za-z0-9_]+$")


@protected.get("/radar/tile/{z}/{x}/{y}")
async def radar_tile(z: int, x: int, y: int, path: str) -> Response:
    import httpx

    # `path` comes from the meta payload (e.g. "/v2/radar/1650000000" or
    # "/v2/radar/nowcast_a1b2c3"); validate strictly — this is a proxy.
    if not _RADAR_PATH_RE.match(path):
        raise HTTPException(422, "invalid radar frame path")
    if not (2 <= z <= 12):
        raise HTTPException(422, "zoom out of range")
    key = f"{path}/{z}/{x}/{y}"
    if key in _radar_tiles:
        _radar_tiles.move_to_end(key)
        return Response(_radar_tiles[key], media_type="image/png",
                        headers={"Cache-Control": "public, max-age=600"})
    # color scheme 2 (universal blue), smoothed, snow shown
    url = f"https://tilecache.rainviewer.com{path}/256/{z}/{x}/{y}/2/1_1.png"
    try:
        async with httpx.AsyncClient(timeout=8) as client:
            r = await client.get(url)
            r.raise_for_status()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(502, f"radar tile fetch failed: {exc}") from exc
    _radar_tiles[key] = r.content
    while len(_radar_tiles) > _RADAR_TILE_CACHE:
        _radar_tiles.popitem(last=False)
    return Response(r.content, media_type="image/png",
                    headers={"Cache-Control": "public, max-age=600"})


# -- sensor placements -------------------------------------------------------
@protected.get("/placements", response_model=list[PlacementOut])
async def list_placements(session: AsyncSession = Depends(get_session)) -> list[models.SensorPlacement]:
    result = await session.execute(select(models.SensorPlacement).order_by(models.SensorPlacement.entity_id))
    return list(result.scalars())


@protected.put("/placements/{entity_id}", response_model=PlacementOut)
async def upsert_placement(
    entity_id: str, body: PlacementIn, session: AsyncSession = Depends(get_session)
) -> models.SensorPlacement:
    result = await session.execute(
        select(models.SensorPlacement).where(models.SensorPlacement.entity_id == entity_id)
    )
    placement = result.scalar_one_or_none()
    if placement is None:
        placement = models.SensorPlacement(entity_id=entity_id)
        session.add(placement)
    placement.room = body.room
    placement.floor = body.floor
    placement.x = body.x
    placement.y = body.y
    placement.icon = body.icon
    await session.commit()
    await session.refresh(placement)
    return placement


@protected.delete("/placements/{entity_id}", status_code=204)
async def delete_placement(entity_id: str, session: AsyncSession = Depends(get_session)) -> None:
    result = await session.execute(
        select(models.SensorPlacement).where(models.SensorPlacement.entity_id == entity_id)
    )
    placement = result.scalar_one_or_none()
    if placement is None:
        raise HTTPException(404, "no placement for that entity")
    await session.delete(placement)
    await session.commit()


# -- panel layouts -----------------------------------------------------------
@protected.get("/layouts/{panel_key}", response_model=LayoutOut)
async def get_layout(panel_key: str, session: AsyncSession = Depends(get_session)) -> models.PanelLayout:
    result = await session.execute(
        select(models.PanelLayout).where(
            models.PanelLayout.panel_key == panel_key, models.PanelLayout.user_id.is_(None)
        )
    )
    layout = result.scalar_one_or_none()
    if layout is None:
        raise HTTPException(404, f"no layout saved for panel: {panel_key}")
    return layout


@protected.put("/layouts/{panel_key}", response_model=LayoutOut)
async def put_layout(
    panel_key: str, body: LayoutIn, session: AsyncSession = Depends(get_session)
) -> models.PanelLayout:
    result = await session.execute(
        select(models.PanelLayout).where(
            models.PanelLayout.panel_key == panel_key, models.PanelLayout.user_id.is_(None)
        )
    )
    layout = result.scalar_one_or_none()
    if layout is None:
        layout = models.PanelLayout(panel_key=panel_key)
        session.add(layout)
    layout.layout_json = body.layout_json
    await session.commit()
    await session.refresh(layout)
    return layout
